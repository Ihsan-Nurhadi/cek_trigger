import sys
import os
import time
import re
import logging
import json
from datetime import datetime
from lxml import etree
from onvif import ONVIFCamera
from zeep import Client
from zeep.wsse.username import UsernameToken
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# Konfigurasi logging yang lebih bersih
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Matikan debug zeep agar tidak berisik
logging.getLogger('zeep').setLevel(logging.WARNING)

# --- KONFIGURASI ---
IP = "192.168.0.107"
PORT = 2020
USER = "nayakapratama"
PASS = "nayakapratama"

# Konfigurasi Excel
EXCEL_FILE = "tapo_detection_log.xlsx"
DETECTION_THRESHOLD = 2.0  # Detik - minimal durasi untuk dicatat


class PeopleDetectionLogger:
    """Class untuk logging deteksi people ke Excel dengan debouncing"""
    
    def __init__(self, excel_file, threshold_seconds=2.0):
        self.excel_file = excel_file
        self.threshold = threshold_seconds
        self.detection_start = None
        self.is_detecting = False
        self.last_detection_time = None
        self.workbook = None
        self.sheet = None
        self.log_count = 0
        
        self.init_excel()
    
    def init_excel(self):
        """Inisialisasi file Excel"""
        if os.path.exists(self.excel_file):
            # Buka file yang sudah ada
            self.workbook = openpyxl.load_workbook(self.excel_file)
            self.sheet = self.workbook.active
            # Hitung jumlah log yang sudah ada
            self.log_count = self.sheet.max_row - 1
            print(f"📂 Excel file ditemukan: {self.excel_file}")
            print(f"📊 Existing logs: {self.log_count}")
        else:
            # Buat file baru
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Detection Log"
            
            # Header
            headers = ["No", "Tanggal", "Waktu Mulai", "Waktu Selesai", 
                      "Durasi (detik)", "Status", "Rule", "Video Source"]
            self.sheet.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in self.sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths
            self.sheet.column_dimensions['A'].width = 8
            self.sheet.column_dimensions['B'].width = 15
            self.sheet.column_dimensions['C'].width = 12
            self.sheet.column_dimensions['D'].width = 12
            self.sheet.column_dimensions['E'].width = 15
            self.sheet.column_dimensions['F'].width = 15
            self.sheet.column_dimensions['G'].width = 25
            self.sheet.column_dimensions['H'].width = 20
            
            self.workbook.save(self.excel_file)
            print(f"✅ Excel file dibuat: {self.excel_file}")
    
    def process_event(self, event_data):
        """Process event dengan debouncing logic"""
        is_detected = event_data.get('is_people_detected', False)
        current_time = datetime.now()
        
        if is_detected:
            if not self.is_detecting:
                # Mulai deteksi baru
                self.detection_start = current_time
                self.is_detecting = True
                print(f"⏱️  Deteksi dimulai: {current_time.strftime('%H:%M:%S')}")
            
            # Update last detection time
            self.last_detection_time = current_time
            
        else:
            # Orang tidak terdeteksi
            if self.is_detecting:
                # Cek apakah ada gap lebih dari 1 detik (berarti orang sudah pergi)
                if self.last_detection_time and (current_time - self.last_detection_time).total_seconds() > 1.0:
                    # Hitung durasi total
                    duration = (self.last_detection_time - self.detection_start).total_seconds()
                    
                    # Jika durasi >= threshold, catat ke Excel
                    if duration >= self.threshold:
                        self.log_to_excel(event_data, duration)
                    else:
                        print(f"⚠️  Deteksi terlalu singkat ({duration:.1f}s), tidak dicatat")
                    
                    # Reset state
                    self.is_detecting = False
                    self.detection_start = None
                    self.last_detection_time = None
    
    def log_to_excel(self, event_data, duration):
        """Catat deteksi ke Excel"""
        self.log_count += 1
        
        # Data untuk Excel
        tanggal = self.detection_start.strftime('%Y-%m-%d')
        waktu_mulai = self.detection_start.strftime('%H:%M:%S')
        waktu_selesai = self.last_detection_time.strftime('%H:%M:%S')
        status = "✅ Tercatat"
        rule = event_data.get('source', {}).get('Rule', '-')
        video_source = event_data.get('source', {}).get('VideoSourceConfigurationToken', '-')
        
        row_data = [
            self.log_count,
            tanggal,
            waktu_mulai,
            waktu_selesai,
            round(duration, 2),
            status,
            rule,
            video_source
        ]
        
        self.sheet.append(row_data)
        
        # Style untuk baris baru
        row_num = self.sheet.max_row
        for cell in self.sheet[row_num]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Highlight durasi jika lama
        duration_cell = self.sheet.cell(row_num, 5)
        if duration > 10:
            duration_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Save file
        self.workbook.save(self.excel_file)
        
        print(f"💾 LOG #{self.log_count} disimpan ke Excel!")
        print(f"   📅 {tanggal} | ⏰ {waktu_mulai} - {waktu_selesai} | ⌛ {duration:.2f}s")
    
    def close(self):
        """Tutup Excel file"""
        if self.workbook:
            # Cek apakah masih ada deteksi yang belum tercatat
            if self.is_detecting and self.last_detection_time:
                current_time = datetime.now()
                duration = (self.last_detection_time - self.detection_start).total_seconds()
                if duration >= self.threshold:
                    print(f"\n⚠️  Menyimpan deteksi terakhir sebelum close...")
                    # Buat dummy event_data
                    dummy_event = {'source': {}, 'is_people_detected': True}
                    self.log_to_excel(dummy_event, duration)
            
            self.workbook.close()
            print(f"\n📊 Total deteksi tercatat: {self.log_count}")
            print(f"📂 File Excel: {self.excel_file}")


def zeep_to_dict(obj, include_private=True):
    """Konversi objek Zeep ke dictionary"""
    if obj is None:
        return None
    
    if isinstance(obj, (str, int, float, bool)):
        return obj
    
    if isinstance(obj, datetime):
        return obj.isoformat()
    
    if isinstance(obj, list):
        return [zeep_to_dict(item, include_private) for item in obj]
    
    if hasattr(obj, '__dict__'):
        result = {}
        for key, value in obj.__dict__.items():
            # Include private attributes jika diminta
            if include_private or not key.startswith('_'):
                result[key] = zeep_to_dict(value, include_private)
        return result
    
    return str(obj)


def parse_xml_element(element):
    """Parse lxml Element menjadi dictionary"""
    result = {}
    
    # Ambil attributes
    for key, value in element.attrib.items():
        # Hilangkan namespace prefix
        key_name = key.split('}')[-1] if '}' in key else key
        result[key_name] = value
    
    # Ambil child elements
    children = {}
    for child in element:
        # Hilangkan namespace dari tag
        tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        
        # Jika child punya children, parse rekursif
        if len(child) > 0:
            child_data = parse_xml_element(child)
            if tag in children:
                # Jika sudah ada, buat jadi list
                if not isinstance(children[tag], list):
                    children[tag] = [children[tag]]
                children[tag].append(child_data)
            else:
                children[tag] = child_data
        else:
            # Leaf node, ambil text atau attributes
            if child.text and child.text.strip():
                value = child.text.strip()
            else:
                value = dict(child.attrib)
            
            if tag in children:
                if not isinstance(children[tag], list):
                    children[tag] = [children[tag]]
                children[tag].append(value)
            else:
                children[tag] = value
    
    result.update(children)
    
    # Jika element punya text dan tidak punya children
    if element.text and element.text.strip() and len(element) == 0:
        return element.text.strip()
    
    return result if result else None


def extract_simple_items_from_xml(element):
    """Ekstrak SimpleItem dari XML Element"""
    result = {}
    
    # Namespace ONVIF
    namespaces = {
        'tt': 'http://www.onvif.org/ver10/schema'
    }
    
    # Cari semua SimpleItem
    simple_items = element.findall('.//tt:SimpleItem', namespaces)
    
    for item in simple_items:
        name = item.get('Name')
        value = item.get('Value')
        if name:
            result[name] = value
    
    return result


def parse_tapo_event(message):
    """Parse event message dari Tapo dan ekstrak informasi penting"""
    try:
        event_data = {
            'timestamp': None,
            'topic': None,
            'event_type': None,
            'is_people_detected': False,
            'source': {},
            'data': {},
            'property_operation': None
        }
        
        # Ekstrak Topic
        if hasattr(message, 'Topic'):
            topic_obj = message.Topic
            if hasattr(topic_obj, '_value_1'):
                event_data['topic'] = str(topic_obj._value_1)
            else:
                event_data['topic'] = str(topic_obj)
            
            # Deteksi tipe event berdasarkan topic
            if 'PeopleDetector/People' in event_data['topic']:
                event_data['event_type'] = 'People Detection'
            elif 'MotionDetector' in event_data['topic']:
                event_data['event_type'] = 'Motion Detection'
            elif 'CellMotionDetector' in event_data['topic']:
                event_data['event_type'] = 'Cell Motion Detection'
            else:
                event_data['event_type'] = 'Unknown Event'
        
        # Ekstrak Message content - INI ADALAH XML ELEMENT!
        if hasattr(message, 'Message'):
            msg_obj = message.Message
            
            # Akses _value_1 yang adalah lxml Element
            if hasattr(msg_obj, '_value_1') and msg_obj._value_1 is not None:
                xml_element = msg_obj._value_1
                
                # Namespace ONVIF
                ns = {'tt': 'http://www.onvif.org/ver10/schema'}
                
                # Ekstrak PropertyOperation
                prop_op = xml_element.get('PropertyOperation')
                if prop_op:
                    event_data['property_operation'] = prop_op
                
                # Ekstrak UtcTime
                utc_time = xml_element.get('UtcTime')
                if utc_time:
                    event_data['timestamp'] = utc_time
                
                # Ekstrak Source
                source_elem = xml_element.find('tt:Source', ns)
                if source_elem is not None:
                    event_data['source'] = extract_simple_items_from_xml(source_elem)
                
                # Ekstrak Data
                data_elem = xml_element.find('tt:Data', ns)
                if data_elem is not None:
                    event_data['data'] = extract_simple_items_from_xml(data_elem)
                    
                    # Cek apakah terdeteksi people
                    if event_data['data'].get('IsPeople') == 'true':
                        event_data['is_people_detected'] = True
                    elif event_data['data'].get('State') == 'true':
                        event_data['is_people_detected'] = True
        
        return event_data
        
    except Exception as e:
        logging.error(f"Error parsing event: {e}")
        import traceback
        traceback.print_exc()
        return {
            'error': str(e), 
            'message': 'Failed to parse event'
        }


def get_tapo_proof_data():
    print("\n" + "="*60)
    print("🎥 TAPO EVENT MONITOR - XML to JSON with Excel Logging")
    print("="*60)
    print(f"📡 Kamera: {IP}:{PORT}")
    print(f"👤 User: {USER}")
    print(f"⏱️  Detection Threshold: {DETECTION_THRESHOLD} detik")
    print(f"📊 Excel File: {EXCEL_FILE}")
    print("="*60 + "\n")
    
    current_dir = os.path.dirname(os.path.abspath(__file__))
    wsdl_path = os.path.join(current_dir, 'wsdl')
    
    if not os.path.exists(wsdl_path):
        print(f"❌ [ERROR] Folder wsdl tidak ditemukan di {wsdl_path}!")
        return

    # Inisialisasi logger
    logger = PeopleDetectionLogger(EXCEL_FILE, DETECTION_THRESHOLD)

    try:
        # 1. KONEKSI AWAL
        print(f"🔌 Menghubungkan ke kamera...")
        cam = ONVIFCamera(IP, PORT, USER, PASS, wsdl_dir=wsdl_path)
        
        # 2. SUBSCRIBE
        print("📝 Membuat subscription...")
        event_service = cam.create_events_service()
        subscription_response = event_service.CreatePullPointSubscription()
        
        # 3. PERBAIKI ALAMAT
        try:
            raw_url = subscription_response.SubscriptionReference.Address._value_1
        except:
            raw_url = subscription_response.SubscriptionReference.Address

        final_url = raw_url
        if f":{PORT}/" not in raw_url:
            final_url = re.sub(r':\d+/', f':{PORT}/', raw_url)
            print(f"🔧 URL diperbaiki: {final_url}")

        # 4. BUAT KONEKSI SECURE
        print("🔐 Membuat koneksi secure (WSSE)...")
        events_wsdl_file = os.path.join(wsdl_path, 'events.wsdl')
        binding_name = '{http://www.onvif.org/ver10/events/wsdl}PullPointSubscriptionBinding'
        
        token = UsernameToken(USER, PASS, use_digest=True)
        pullpoint_client = Client(wsdl=events_wsdl_file, transport=cam.transport, wsse=token)
        pullpoint = pullpoint_client.create_service(binding_name, final_url)

        print("\n" + "="*60)
        print("✅ LISTENER AKTIF - Menunggu event dari kamera...")
        print("👋 Gerakkan tangan di depan kamera untuk trigger event!")
        print(f"📝 Deteksi akan dicatat jika berlangsung minimal {DETECTION_THRESHOLD} detik")
        print("="*60 + "\n")

        event_counter = 0

        # 5. LOOPING PENGAMBILAN DATA
        while True:
            try:
                response = pullpoint.PullMessages(Timeout='PT5S', MessageLimit=10)
                
                # Cek apakah ada NotificationMessage
                if hasattr(response, 'NotificationMessage'):
                    messages = response.NotificationMessage
                    
                    # Pastikan messages adalah list
                    if not isinstance(messages, list):
                        messages = [messages]
                    
                    for msg in messages:
                        event_counter += 1
                        
                        # Parse event ke JSON
                        event_json = parse_tapo_event(msg)
                        
                        # Proses untuk logging dengan debouncing
                        logger.process_event(event_json)
                        
                        # Tampilkan dengan format yang bagus
                        print("\n" + "─"*60)
                        print(f"📋 EVENT #{event_counter} - {datetime.now().strftime('%H:%M:%S')}")
                        print("─"*60)
                        
                        # Tampilkan ringkasan dengan warna
                        if event_json.get('is_people_detected'):
                            print("🚨 STATUS: PEOPLE DETECTED! 🚨")
                        else:
                            print(f"📌 Status: No People")
                        
                        print(f"📂 Event Type: {event_json.get('event_type', 'Unknown')}")
                        print(f"⏰ Timestamp: {event_json.get('timestamp', 'N/A')}")
                        print(f"📡 Topic: {event_json.get('topic', 'N/A')}")
                        print(f"🔄 Operation: {event_json.get('property_operation', 'N/A')}")
                        
                        # Tampilkan Source
                        if event_json.get('source'):
                            print(f"\n📍 SOURCE:")
                            for key, val in event_json['source'].items():
                                print(f"   • {key}: {val}")
                        
                        # Tampilkan Data
                        if event_json.get('data'):
                            print(f"\n📊 DATA:")
                            for key, val in event_json['data'].items():
                                icon = "✅" if val == "true" else "❌"
                                print(f"   {icon} {key}: {val}")
                        
                        # Tampilkan JSON lengkap (opsional, uncomment jika perlu)
                        # print(f"\n📄 FULL JSON:")
                        # json_str = json.dumps(event_json, indent=2, ensure_ascii=False)
                        # print(json_str)
                        print("─"*60)
                
                # Juga proses event kosong untuk update debouncing state
                else:
                    dummy_event = {'is_people_detected': False}
                    logger.process_event(dummy_event)
                
                time.sleep(0.1)
                
            except Exception as e:
                error_msg = str(e).lower()
                if "timeout" not in error_msg and "no messages" not in error_msg:
                    logging.warning(f"⚠️  Error saat pulling messages: {e}")
                time.sleep(1)

    except KeyboardInterrupt:
        print("\n\n" + "="*60)
        print("👋 Program dihentikan oleh user.")
        print(f"📊 Total events yang ditangkap: {event_counter}")
        logger.close()
        print("="*60)
        
    except Exception as e:
        print(f"\n❌ [FATAL ERROR] {e}")
        import traceback
        traceback.print_exc()
        logger.close()


if __name__ == "__main__":
    get_tapo_proof_data()